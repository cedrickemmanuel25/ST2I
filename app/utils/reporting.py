import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime
from typing import List

def generate_excel_report(data: List[dict]):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Présences"

    # Header styling
    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    
    headers = ["Date", "Utilisateur", "Type", "Heure", "Statut", "Note"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row in data:
        ws.append([
            row["timestamp"].split("T")[0],
            f"{row['user']['prenom']} {row['user']['nom']}",
            row["type"].upper(),
            row["timestamp"].split("T")[1].split(".")[0],
            row["statut"].upper(),
            row.get("note", "")
        ])

    # Auto-adjust column width
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    wb.save(output)
    output.seek(0)
    return output

def generate_pdf_report(data: List[dict]):
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    elements.append(Paragraph("Rapport de Présences ST2I", styles['Title']))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Généré le : {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 24))

    # Table Data
    table_data = [["Date", "Utilisateur", "Type", "Heure", "Statut"]]
    for row in data:
        table_data.append([
            row["timestamp"].split("T")[0],
            f"{row['user']['prenom']} {row['user']['nom']}",
            row["type"].upper(),
            row["timestamp"].split("T")[1].split(".")[0],
            row["statut"]
        ])

    t = Table(table_data, colWidths=[80, 150, 60, 80, 100])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    
    elements.append(t)
    doc.build(elements)
    output.seek(0)
    return output
